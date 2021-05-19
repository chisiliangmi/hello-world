import openpyxl
import re


class CalNum:
    def __init__(self, file_path):
        self.wb = openpyxl.load_workbook(file_path)
        self.sheet_names = self.wb.get_sheet_names()

    def run_cal_num(self, city):
        # 使用正则表达式以匹配工作表名字
        regex = ['.*去', '.*回', '\\W', '^X8.*']
        global symbol
        symbol = []
        for i in range(len(self.sheet_names)):
            for j in range(len(regex)):
                Regex = re.compile(regex[j])
                reg = Regex.search(self.sheet_names[i])
                if reg and j == 0:
                    symbol1 = i
                    global num_leave_city
                    num_leave_city = self.zhida_leave(
                        symbol1, city)
                    print("出境 %s 技术直达： %d" % (city, num_leave_city))
                elif reg and j == 1:
                    symbol_2 = i
                    global num_back_city
                    num_back_city = self.zhida_back(
                        symbol_2, city)
                    print("回程 %s 技术直达： %d" % (city, num_back_city))
                    print('\n')
                elif reg and j == 2:
                    symbol.append(0)
                elif reg and j == 3:
                    symbol.append(i)

    # 技术直达车次
    # 计算去程
    def zhida_leave(self, symbol1, city):
        # 要操作的表
        sheet = self.wb[self.sheet_names[symbol1]]
        num_leave_city_ = 0
        for row in range(5, sheet.max_row + 1):
            # 确定列车到达
            if sheet['Y' + str(row)].value == city and sheet[
                    'D' + str(row)].value is not None:
                num_leave_city_ += 1

        return num_leave_city_


    # 计算回程
    def zhida_back(self, symbol2, city):
        sheet = self.wb[self.sheet_names[symbol2]]
        num_back_city_ = 0
        for row in range(5, sheet.max_row + 1):
            # 确定列车出发
            if sheet['C' + str(row)].value == str(city) and sheet[
                    'D' + str(row)].value is not None:
                num_back_city_ += 1

        return num_back_city_

    def run_cal_num_(self, city):
        sym = []
        for i in range(len(symbol)):
            if symbol[i] == 0:
                sym.append(i)
                continue
        sym.insert(0,0)
        for i in range(len(sym)):
            if i == 0:
                symb = symbol[:sym[i+1]]
                self.banlie(symb, city)
            elif i == len(sym)-1:
                symb = symbol[sym[i]+1:]
                self.banlie(symb, city)
            else:
                symb = symbol[sym[i]+1:sym[i+1]]
                self.banlie(symb, city)

    def check_train_leave(sheet, city):
        if sheet['S3'].value == city:
            return True
        else:
            return False

    def check_train_back(sheet, city):
        if sheet['C3'].value == city:
            return True
        else:
            return False

    def banlie(self, symb, city):
        number_leave_city = 0
        number_back_city = 0
        for i in range(len(symb)):
            sheet = self.wb[self.sheet_names[symb[i]]]
            if CalNum.check_train_leave(sheet, city):
                number_leave_city_ = CalNum.train_leave(sheet, city)
                number_leave_city += number_leave_city_
            elif CalNum.check_train_back(sheet, city):
                number_back_city_ = CalNum.train_back(sheet, city)
                number_back_city += number_back_city_
        if number_leave_city != 0:
            print("出境 %s 班列： %d" % (city, number_leave_city))
        elif number_back_city != 0:
            print("回程 %s 班列： %d" % (city, number_back_city))
            print("\n")

    # 计算一个工作表的班列车次
    # 计算去程
    def train_leave(sheet, city):
        number_leave_city_ = 0
        for row in range(10, 60):
            if CalNum.check_train_leave(sheet, city) and sheet[
                    'B' + str(row)].value is not None:
                number_leave_city_ += 1

        return number_leave_city_

    # 计算回程
    def train_back(sheet, city):
        number_back_city_ = 0
        for row in range(10, 60):
            if CalNum.check_train_back(sheet, city) and sheet[
                    'B' + str(row)].value is not None:
                number_back_city_ += 1

        return number_back_city_
